import jdatetime as jdatetime
from django_filters.rest_framework import DjangoFilterBackend
from drf_multiple_model.views import ObjectMultipleModelAPIView
from rest_framework.filters import SearchFilter
from rest_framework.generics import ListAPIView, RetrieveAPIView, CreateAPIView, get_object_or_404
from rest_framework.parsers import MultiPartParser, FormParser, FileUploadParser
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.writer.excel import save_virtual_workbook

from django.http import HttpResponse

from core.api_v1.serilalizers import ExcelPatternSerializer, ExcelPatternUploadedFileSerializer, \
    ExcelPatternUploadedFileCreateSerializer
from core.models import ExcelPattern, ExcelPatternUploadedFile


class ExcelPatternListView(ListAPIView):
    permission_classes = (IsAuthenticated,)
    queryset = ExcelPattern.objects.all()
    serializer_class = ExcelPatternSerializer

    def get_queryset(self):
        user_profile = self.request.user.user_profile
        return ExcelPattern.objects.filter(owner=user_profile)


def to_not_none_str(cell_value):
    if (cell_value is not None):
        return str(cell_value)
    return ''


def sample_1_to_json(wb):
    # print('initalizing')
    # wb = load_workbook(filename='sample1.xlsx')
    print('excel loaded')
    # grab the active worksheet
    ws = wb.active
    print('active worksheet loaded')
    transactions = []
    i = 2
    while i < 625:
        ii = str(i)
        transaction_amount = int(to_not_none_str(ws['B' + ii].value))
        if transaction_amount < 0:
            i += 1
            continue
        transactions.append({
            'date': to_not_none_str(ws['L' + ii].value),
            'time': to_not_none_str(ws['K' + ii].value),
            'withdraw': '0',
            'deposit': str(transaction_amount),
            'balance': to_not_none_str(ws['A' + ii].value),
        })
        i += 1
    return transactions


def str_to_jdate2(date_str):
    year = int(date_str.split('/')[0])
    month = int(date_str.split('/')[1])
    day = int(date_str.split('/')[2])
    return jdatetime.datetime(year, month, day)


def sort_key(transaction):
    return str_to_jdate2(transaction['date'])


def generate_excel(from_wb):
    sample_1_json = sample_1_to_json(from_wb)
    transactions = sample_1_json[::]
    transactions.sort(key=sort_key)
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'تاریخ'
    ws['B1'] = 'ساعت'
    ws['C1'] = 'مبلغ'
    ws['D1'] = 'مانده'
    i = 2
    j = 0
    while j < len(transactions):
        ws['A' + str(i)] = transactions[j]['date']
        ws['B' + str(i)] = transactions[j]['time']
        ws['C' + str(i)] = transactions[j]['deposit']
        ws['D' + str(i)] = transactions[j]['balance']
        i += 1
        j += 1
    rows = range(1, i)
    columns = range(1, 15)
    for row in rows:
        for col in columns:
            ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')

    # wb.save('sample1-res.xlsx')
    return wb

class ExcelPatternUploadedFileCreateView(CreateAPIView):
    # TODO: fix authentication
    # permission_classes = (IsAuthenticated,)
    permission_classes = (AllowAny,)
    queryset = ExcelPatternUploadedFile.objects.all()
    serializer_class = ExcelPatternUploadedFileCreateSerializer
    parser_classes = (FormParser, MultiPartParser, FileUploadParser)

    def create(self, request, *args, **kwargs):
        # TODO: only use the pattern for the owner
        # user_profile = self.request.user.user_profile
        # the_pattern = get_object_or_404(ExcelPattern.objects.all(), owner=user_profile, pk=kwargs['pk'])

        the_pattern = get_object_or_404(ExcelPattern.objects.all(), pk=kwargs['pk'])
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        # serializer.save(pattern=the_pattern)

        from_wb = load_workbook(filename=request.FILES['the_file'].file)
        workbook = generate_excel(from_wb)

        response = HttpResponse(content=save_virtual_workbook(workbook),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=myexport.xlsx'
        return response
